"""Functions for generating and dispatching alerts."""

from __future__ import annotations

import io
import json
import threading
import time
from datetime import datetime, timedelta
from pathlib import Path

from loguru import logger
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from redis.exceptions import RedisError

from core import events
from modules.profiler import register_thread

from .email_utils import send_email


# AlertWorker class encapsulates alertworker behavior
class AlertWorker:
    # __init__ routine
    def __init__(self, cfg: dict, redis_client, base_dir: Path):
        """Create a worker and launch the alert processing thread."""
        self.cfg = cfg
        self.redis = redis_client
        # Retention window for Redis keys set by this worker
        self.retention_secs = int(cfg.get("alert_key_retention_secs", 7 * 24 * 60 * 60))
        self.base_dir = base_dir
        self.running = True
        self.thread = threading.Thread(target=self.loop, daemon=True)
        self.thread.start()

    # stop routine
    def stop(self):
        """Signal the worker thread to stop and wait briefly for it."""
        self.running = False
        self.thread.join(timeout=2)

    # loop routine
    def loop(self):
        """Main worker loop that periodically evaluates alert rules."""
        register_thread("Alerts")
        logger.info("AlertWorker started")
        while self.running:
            start = time.monotonic()
            try:
                self.check_rules()
                self.check_overdue_gatepasses()
            except (RuntimeError, RedisError, ValueError) as exc:
                logger.exception("alert loop error: {}", exc)
            elapsed = time.monotonic() - start
            self._log_cycle(elapsed)
            time.sleep(60)
        logger.info("AlertWorker stopped")

    def _log_cycle(self, elapsed: float) -> None:
        """Log completion of a worker cycle."""
        logger.info(f"AlertWorker cycle completed in {elapsed:.1f}s")

    # check_overdue_gatepasses routine
    def check_overdue_gatepasses(self) -> None:
        """Mark issued gate passes past expiry as overdue and alert."""
        now = int(time.time())
        try:
            entries = self.redis.zrange("vms_logs", 0, -1)
        except RedisError as exc:
            logger.exception("failed to scan gate passes: {}", exc)
            return
        for e in entries:
            try:
                obj = json.loads(e if isinstance(e, str) else e.decode())
            except json.JSONDecodeError:
                continue
            if obj.get("status") not in {"Issued", "approved"}:
                continue
            if int(obj.get("valid_to", now)) >= now:
                continue
            obj["status"] = "Overdue"
            self.redis.zrem("vms_logs", e)
            self.redis.zadd("vms_logs", {json.dumps(obj): obj["ts"]})
            try:
                self.redis.hset(f"gatepass:pass:{obj['gate_id']}", "status", "Overdue")
            except RedisError:
                pass
            recipient = self.cfg.get("security_email")
            if recipient:
                send_email(
                    "Gate pass overdue",
                    f"Gate pass {obj['gate_id']} for {obj.get('name','')} is overdue",
                    [recipient],
                    self.cfg.get("email", {}),
                )

    # _collect_rows routine
    def _collect_rows(self, key: str, start_ts: int, end_ts: int, filter_fn=None):
        """Return decoded rows from ``key`` within the time window."""
        entries = self.redis.zrangebyscore(key, start_ts + 1, end_ts)
        rows = []
        for item in entries:
            try:
                e = json.loads(item if isinstance(item, str) else item.decode())
            except json.JSONDecodeError:
                continue
            if filter_fn and not filter_fn(e):
                continue
            rows.append(e)
        return rows

    # _send_vms_report routine
    def _send_vms_report(self, rows, recipients, subject, attach=True):
        """Email a visitor management report built from recent log rows."""
        wb = Workbook()
        ws = wb.active
        ws.append(["Time", "Gate", "Name", "Phone", "Host"])
        for r in rows:
            ws.append(
                [
                    datetime.fromtimestamp(r["ts"]).strftime("%Y-%m-%d %H:%M"),
                    r.get("gate_id"),
                    r.get("name"),
                    r.get("phone"),
                    r.get("host"),
                ]
            )
        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        send_email(
            subject,
            "See attached visitor report",
            recipients,
            self.cfg.get("email", {}),
            attachment=bio.getvalue(),
            attachment_name="vms_report.xlsx",
            attachment_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    # _send_report routine
    def _send_report(self, rows, recipients, subject, attach=True):
        """Compile PPE log rows into a spreadsheet and email it."""
        wb = Workbook()
        ws = wb.active
        ws.append(["Time", "Camera", "Track", "Status", "Conf", "Color"])
        for r in rows:
            ws.append(
                [
                    datetime.fromtimestamp(r["ts"]).strftime("%Y-%m-%d %H:%M"),
                    r.get("cam_id"),
                    r.get("track_id"),
                    r.get("status"),
                    round(r.get("conf", 0), 2),
                    r.get("color") or "",
                ]
            )
            path = r.get("path")
            if path and Path(path).exists():
                img = XLImage(path)
                img.width = 80
                img.height = 60
                ws.add_image(img, f"F{ws.max_row}")
        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        attachment = bio.getvalue() if attach else None
        send_email(
            subject,
            "See attached report" if attach else "Alert",
            recipients,
            self.cfg.get("email", {}),
            attachment=attachment,
            attachment_name="report.xlsx" if attach else None,
            attachment_type=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                if attach
                else None
            ),
        )

    # check_rules routine
    def check_rules(self):
        """Evaluate configured alert rules and send notifications as needed."""
        if not self.cfg.get("email_enabled", False):
            return
        rules = self.cfg.get("alert_rules", [])
        if not rules:
            return
        now = int(time.time())
        for i, rule in enumerate(rules):
            metric = rule.get("metric")
            rtype = rule.get("type", "event")
            value = int(rule.get("value", 1))
            attach = rule.get("attach", True)
            recipients = [
                a.strip() for a in rule.get("recipients", "").split(",") if a.strip()
            ]
            if not metric or not recipients:
                continue
            last_key = f"alert_rule_{i}_last"
            prev_key = f"alert_rule_{i}_pending" if rtype == "threshold" else None
            pipe = self.redis.pipeline()
            pipe.get(last_key)
            if prev_key:
                pipe.get(prev_key)
            res = pipe.execute()
            last_ts = int(float(res[0] or 0))
            pending = int(res[1] or 0) if prev_key else 0
            if metric == events.VISITOR_REGISTERED:
                fetch_rows = lambda s, e: self._collect_rows("vms_logs", s, e)
                send_report = self._send_vms_report
            elif metric in events.ALL_EVENTS:
                fetch_rows = lambda s, e: self._collect_rows(
                    "events", s, e, lambda r: r.get("event") == metric
                )
                send_report = self._send_report
            else:
                fetch_rows = lambda s, e: self._collect_rows(
                    "ppe_logs", s, e, lambda r: r.get("status") == metric
                )
                send_report = self._send_report
            if rtype == "frequency":
                interval = value * 60
                if now - last_ts >= interval:
                    rows = fetch_rows(last_ts, now)
                    if rows:
                        send_report(rows, recipients, f"Alert: {metric}", attach)
                    self.redis.set(last_key, now)
                    self.redis.expire(last_key, self.retention_secs)
                continue

            rows = fetch_rows(last_ts, now)
            if not rows:
                continue
            if rtype == "event":
                if len(rows) >= value:
                    send_rows = rows[:value]
                    send_report(send_rows, recipients, f"Alert: {metric}", attach)
                    self.redis.set(last_key, send_rows[-1]["ts"])
                    self.redis.expire(last_key, self.retention_secs)
            elif rtype == "threshold":
                pending += len(rows)
                pipe = self.redis.pipeline()
                if pending >= value:
                    send_rows = rows[:value]
                    send_report(send_rows, recipients, f"Alert: {metric}", attach)
                    pending -= value
                    self.redis.set(last_key, send_rows[-1]["ts"])
                    self.redis.expire(last_key, self.retention_secs)
                self.redis.set(prev_key, pending)
                self.redis.expire(prev_key, self.retention_secs)
